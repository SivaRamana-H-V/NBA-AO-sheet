import openpyxl
import streamlit as st

def ReadValue():
    
    uploaded_file = st.file_uploader("Upload Excel Files IA2 Mark",type=['xlsx','csv'])
    if uploaded_file is not None:
         path = uploaded_file.name #"D:\\Users\\hp\\Documents\\XL\\IA1_CSE_1620_CY6151.xlsx"
         wb_obj=openpyxl.load_workbook(path)
         sheet_obj=wb_obj.active
         max_row = sheet_obj.max_row
         CO3=[]
         for i in range(6,max_row+1):
            cell_obj=sheet_obj.cell(row=i,column=12)    
            CO3.append(cell_obj.value)
         CO4=[]
         for i in range(6,max_row+1):
            cell_obj=sheet_obj.cell(row=i,column=13)    
            CO4.append(cell_obj.value)
         
         return CO3,CO4 