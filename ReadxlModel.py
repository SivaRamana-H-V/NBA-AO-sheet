import openpyxl
import streamlit as st

def ReadValue():
    
    uploaded_file = st.file_uploader("Upload Excel Files Model Mark",type=['xlsx','csv','xls'])
    if uploaded_file is not None:
         path = uploaded_file #"D:\\Users\\hp\\Documents\\XL\\IA1_CSE_1620_CY6151.xlsx"
         wb_obj=openpyxl.load_workbook(path)
         sheet_obj=wb_obj.active
         max_row = sheet_obj.max_row
         CO1=[]
         for i in range(6,max_row+1):
            cell_obj=sheet_obj.cell(row=i,column=17)    
            CO1.append(cell_obj.value)
         CO2=[]
         for i in range(6,max_row+1):
            cell_obj=sheet_obj.cell(row=i,column=18)    
            CO2.append(cell_obj.value)
         CO3=[]
         for i in range(6,max_row+1):
            cell_obj=sheet_obj.cell(row=i,column=19)    
            CO3.append(cell_obj.value)
         CO4=[]
         for i in range(6,max_row+1):
            cell_obj=sheet_obj.cell(row=i,column=20)    
            CO4.append(cell_obj.value)
         CO5=[]
         for i in range(6,max_row+1):
            cell_obj=sheet_obj.cell(row=i,column=21)    
            CO5.append(cell_obj.value)
         
         return CO1,CO2,CO3,CO4,CO5 
