import openpyxl
import streamlit as st

def ReadValue():
    
   uploaded_file = st.file_uploader("Upload Excel Files IA1 Mark",type=['xlsx','csv'])
   if uploaded_file is not None:
      path = uploaded_file #"D:\\Users\\hp\\Documents\\XL\\IA1_CSE_1620_CY6151.xlsx"
      wb_obj=openpyxl.load_workbook(path)
      sheet_obj=wb_obj.active
      max_row = sheet_obj.max_row
      CO1=[]
      for i in range(6,max_row+1):
         cell_obj=sheet_obj.cell(row=i,column=10)    
         CO1.append(cell_obj.value)
      CO2=[]
      for i in range(6,max_row+1):
         cell_obj=sheet_obj.cell(row=i,column=11)    
         CO2.append(cell_obj.value)
         
      return CO1,CO2 
